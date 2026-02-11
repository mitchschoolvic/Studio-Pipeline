"""
Analytics Excel Export Service

Exports analytics data to Excel file for Power Automate integration.
Creates Excel table (not just sheet) for proper Flow compatibility.

Also produces timestamped CSV archives for audit trails.

Only included when BUILD_WITH_AI is enabled.
"""
import csv
import logging
import os
from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from typing import List, Optional

from models_analytics import FileAnalytics

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("⚠️  openpyxl not available - Excel export disabled")

logger = logging.getLogger(__name__)


class AnalyticsExcelService:
    """
    Service that exports analytics to Excel file for Power Automate.
    
    Key Features:
    - Creates Excel TABLE (required for Power Automate to read)
    - Atomic writes (tmp → rename) for OneDrive/SharePoint safety
    - Proper type handling (numbers as numbers, not strings)
    - Also produces timestamped CSV for audit trails
    
    Excel Format matches your specification:
    - Title, Description, Duration, DurationSeconds, Type, Faculty, 
      Speaker, Audience, Timestamp, TimestampSort, ThumbnailUrl,
      Filename, StudioLocation, Language, SpeakerCount, Transcript, VideoUrl
    """
    
    # Excel field order (must match CSV headers)
    EXCEL_HEADERS = [
        'Title',
        'Description',
        'Duration',
        'DurationSeconds',
        'Type',
        'Faculty',
        'Speaker',
        'Audience',
        'Timestamp',
        'TimestampSort',
        'ThumbnailUrl',
        'ThumbnailPath',
        'Filename',
        'StudioLocation',
        'Language',
        'SpeakerCount',
        'Transcript',
        'VideoUrl'
    ]
    
    # Fields that should be numbers (not strings)
    NUMERIC_FIELDS = {'DurationSeconds', 'SpeakerCount'}
    
    def __init__(self, db: Session):
        self.db = db
        
    def get_output_dir(self) -> Path:
        """
        Get directory for analytics exports.
        
        Returns:
            Path to analytics output directory
        """
        from models import Setting
        
        # Get analytics output path from settings
        setting = self.db.query(Setting).filter(
            Setting.key == 'analytics_output_path'
        ).first()
        
        if setting and setting.value:
            output_dir = Path(setting.value)
        else:
            # Default to subdirectory of output path
            output_setting = self.db.query(Setting).filter(
                Setting.key == 'output_path'
            ).first()
            
            if output_setting and output_setting.value:
                output_dir = Path(output_setting.value) / 'analytics'
            else:
                output_dir = Path.home() / 'VideoAnalytics'
        
        # Create directory if it doesn't exist
        output_dir.mkdir(parents=True, exist_ok=True)
        
        return output_dir
    
    def get_excel_path(self) -> Path:
        """
        Get path for the main Excel file (analytics.xlsx).
        
        This is the file that Power Automate reads.
        Always overwritten with latest data.
        
        Returns:
            Path to analytics.xlsx file
        """
        return self.get_output_dir() / 'analytics.xlsx'
    
    def get_csv_archive_path(self) -> Path:
        """
        Get path for timestamped CSV archive.
        
        Returns:
            Path to timestamped CSV file (never overwritten)
        """
        output_dir = self.get_output_dir()
        timestamp = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
        return output_dir / f'analytics-{timestamp}.csv'
    
    async def export_to_excel(self, include_pending: bool = False) -> Path:
        """
        Export all completed analytics to Excel file with Power Automate support.
        
        Process:
        1. Query analytics records
        2. Write to analytics.xlsx.tmp (atomic safety)
        3. Create Excel TABLE (Power Automate requirement)
        4. Flush and fsync
        5. Atomic rename to analytics.xlsx
        6. Also create timestamped CSV archive
        
        Args:
            include_pending: If True, include PENDING/TRANSCRIBING/ANALYZING records
            
        Returns:
            Path to the Excel file
            
        Raises:
            RuntimeError: If openpyxl not available
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed - cannot export to Excel")
        
        # Query completed analytics (eagerly load file and session for VideoURL and thumbnails)
        from sqlalchemy.orm import joinedload
        from models import File
        query = self.db.query(FileAnalytics).options(
            joinedload(FileAnalytics.file).joinedload(File.session)
        )

        if not include_pending:
            query = query.filter(FileAnalytics.state == 'COMPLETED')

        analytics_records = query.order_by(FileAnalytics.created_at.desc()).all()

        if not analytics_records:
            logger.warning("No analytics records to export")
            return None

        # Get paths
        excel_path = self.get_excel_path()
        tmp_path = excel_path.with_suffix('.xlsx.tmp')
        csv_path = self.get_csv_archive_path()

        # Export thumbnails to thumbnails subfolder
        thumbnail_urls = self._export_thumbnails(analytics_records)

        # Write Excel file with thumbnail URLs
        self._write_excel(analytics_records, tmp_path, thumbnail_urls)
        
        # Atomic rename (safe for OneDrive/SharePoint)
        os.replace(str(tmp_path), str(excel_path))
        
        # Also write CSV archive
        self._write_csv(analytics_records, csv_path)
        
        logger.info(f"✅ Exported {len(analytics_records)} analytics records")
        logger.info(f"   Excel: {excel_path}")
        logger.info(f"   CSV Archive: {csv_path}")
        
        return excel_path

    def _export_thumbnails(self, records: List[FileAnalytics]) -> dict:
        """
        Copy thumbnails from system cache to export folder.
        Only copies if destination doesn't exist or source is newer.

        Args:
            records: List of FileAnalytics records with file/session relationships loaded

        Returns:
            Dict mapping file_id → {'url': relative_path, 'path': absolute_path}
        """
        from pathlib import Path
        from urllib.parse import quote
        import shutil

        # Get export directory and create thumbnails subfolder
        output_dir = self.get_output_dir()
        thumbnails_dir = output_dir / 'thumbnails'
        thumbnails_dir.mkdir(parents=True, exist_ok=True)

        thumbnail_urls = {}
        copied_count = 0
        skipped_count = 0

        for record in records:
            # Skip if no file or thumbnail not ready
            if not record.file or record.file.thumbnail_state != 'READY':
                continue

            # Get source thumbnail path from database
            if not record.file.thumbnail_path:
                continue

            source_path = Path(record.file.thumbnail_path)

            # Skip if source doesn't exist
            if not source_path.exists():
                logger.warning(f"Thumbnail not found for file {record.file_id}: {source_path}")
                continue

            # Generate session-based filename
            session = record.file.session
            if not session:
                continue

            # Use session.name directly (already contains studio + date + time)
            session_name = session.name
            thumbnail_ext = source_path.suffix  # Usually .jpg
            dest_filename = f"{session_name}{thumbnail_ext}"
            dest_path = thumbnails_dir / dest_filename

            # Only copy if doesn't exist or source is newer
            should_copy = False
            if not dest_path.exists():
                should_copy = True
            elif source_path.stat().st_mtime > dest_path.stat().st_mtime:
                should_copy = True

            if should_copy:
                try:
                    shutil.copy2(str(source_path), str(dest_path))
                    copied_count += 1
                    logger.debug(f"Copied thumbnail: {dest_filename} ({dest_path.stat().st_size} bytes)")
                except Exception as e:
                    logger.error(f"Failed to copy thumbnail for {record.file_id}: {e}")
                    continue
            else:
                skipped_count += 1

            # Generate URL-encoded relative path for Excel
            relative_path = f"thumbnails/{dest_filename}"
            encoded_path = quote(relative_path, safe='/')
            
            thumbnail_urls[record.file_id] = {
                'url': encoded_path,
                'path': str(dest_path.absolute())
            }

        logger.info(f"Thumbnail export: {copied_count} copied, {skipped_count} skipped (already current)")
        return thumbnail_urls

    def _write_excel(self, records: List[FileAnalytics], path: Path, thumbnail_urls: dict = None):
        """
        Write analytics records to Excel file with TABLE.
        
        Power Automate requires an Excel TABLE (not just a worksheet).
        
        Args:
            records: List of FileAnalytics records
            path: Path to write Excel file
        """
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics"
        
        # Write headers
        for col_idx, header in enumerate(self.EXCEL_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            # Get thumbnail info for this record
            thumb_info = thumbnail_urls.get(record.file_id) if thumbnail_urls else {}
            thumb_url = thumb_info.get('url') if isinstance(thumb_info, dict) else thumb_info
            thumb_path = thumb_info.get('path') if isinstance(thumb_info, dict) else None

            # Generate row data with thumbnail URL and Path
            row_data = record.to_excel_row(
                db_session=self.db, 
                thumbnail_url=thumb_url,
                thumbnail_path=thumb_path
            )

            for col_idx, header in enumerate(self.EXCEL_HEADERS, start=1):
                value = row_data.get(header, '')
                
                # Ensure numeric fields are actual numbers
                if header in self.NUMERIC_FIELDS and value:
                    try:
                        value = int(value) if isinstance(value, (int, float, str)) else value
                    except (ValueError, TypeError):
                        value = 0
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Create Excel TABLE over the data range
        # This is required for Power Automate to properly read the data
        last_row = len(records) + 1
        last_col = len(self.EXCEL_HEADERS)
        last_col_letter = get_column_letter(last_col)
        
        table = Table(
            displayName="AnalyticsTable",
            ref=f"A1:{last_col_letter}{last_row}"
        )
        
        # Add style to table
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        ws.add_table(table)
        
        # Auto-size columns (optional, helps readability)
        for col_idx, header in enumerate(self.EXCEL_HEADERS, start=1):
            col_letter = get_column_letter(col_idx)
            # Set reasonable max width to avoid transcript column being huge
            max_width = 50 if header != 'Transcript' else 100
            ws.column_dimensions[col_letter].width = min(len(header) + 2, max_width)
        
        # Save with flush/fsync for safety
        wb.save(str(path))
        
        # Explicitly flush to disk (important for OneDrive sync)
        with open(path, 'rb') as f:
            os.fsync(f.fileno())
    
    def _write_csv(self, records: List[FileAnalytics], path: Path):
        """
        Write analytics records to CSV file (audit trail).
        
        Args:
            records: List of FileAnalytics records
            path: Path to write CSV file
        """
        with open(path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.EXCEL_HEADERS)
            writer.writeheader()

            for record in records:
                row = record.to_excel_row(db_session=self.db)
                writer.writerow(row)
    
    def export_single_record(self, file_id: str) -> Optional[dict]:
        """
        Export a single analytics record as dictionary.
        
        Args:
            file_id: ID of file to export
            
        Returns:
            Excel row dictionary, or None if not found
        """
        analytics = self.db.query(FileAnalytics).filter(
            FileAnalytics.file_id == file_id
        ).first()

        if not analytics:
            return None

        return analytics.to_excel_row(db_session=self.db)
    
    def get_export_stats(self) -> dict:
        """
        Get statistics about Excel exports.
        
        Returns:
            Dictionary with export statistics
        """
        from sqlalchemy import func
        
        total_completed = self.db.query(func.count(FileAnalytics.id)).filter(
            FileAnalytics.state == 'COMPLETED'
        ).scalar()
        
        total_records = self.db.query(func.count(FileAnalytics.id)).scalar()
        
        excel_path = self.get_excel_path()
        export_exists = excel_path.exists()
        export_size = excel_path.stat().st_size if export_exists else 0
        export_modified = datetime.fromtimestamp(
            excel_path.stat().st_mtime
        ).isoformat() if export_exists else None
        
        # Count CSV archives
        output_dir = self.get_output_dir()
        csv_count = len(list(output_dir.glob('analytics-*.csv')))
        
        return {
            'total_completed': total_completed or 0,
            'total_records': total_records or 0,
            'excel_exists': export_exists,
            'excel_path': str(excel_path),
            'excel_size_bytes': export_size,
            'excel_modified': export_modified,
            'csv_archives': csv_count,
            'output_directory': str(output_dir)
        }
